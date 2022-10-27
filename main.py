import logging
import argparse
import os
import io
import re
import pandas as pd
from openpyxl import load_workbook
import csv
from difflib import SequenceMatcher
import geopandas as gpd
from shapely.geometry import Polygon,Point
from dotenv import load_dotenv
import googlemaps


CALFIRE_SKIPROWS=4
CALFIRE_SKIP_SHEETS = ["Q1"]
NBRSVC_SKIPROWS=0
TREE_NAME_MATCH_RATIO = 0.75

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

logging.basicConfig(format='%(asctime)s %(pathname)s:%(lineno)d - %(name)s - %(levelname)s - %(message)s')

TREE_STANDARD_MODEL_NAMES = {
    "Tree Number":"Tree Number",
    "Funder":"Funder",
    "Scientific Name": "Scientific Name",
    "Species": "Scientific Name",
    "Common Name" : "Common Name",
    "City" : "City",
    "Ownership": "Ownership",
    "X Coordinate": "X Coordinate",
    "Census Tract" : "Census Tract",
    "DAC Status" : "DAC Status",
    "Date Planted" : "Date Planted",
    "Stock Size"  : "Stock Size",
    "Grow Space" : "Grow Space",
    "St Address" : "Street Address",
    "Stret Address" : "Street Address"
}

LB_MIN_X = -118.24892785887
LB_MAX_X = -118.063262203623
LB_MIN_Y =  33.6675353632098
LB_MAX_Y =  33.885547164413

load_dotenv()

lbdf = gpd.read_file('https://services6.arcgis.com/yCArG7wGXGyWLqav/arcgis/rest/services/City_of_Long_Beach_City_Boundary_Official/FeatureServer/0/query?outFields=*&where=1%3D1&f=geojson')
#gdf = gpd.read_file('./colb-boundary.geojson')
LBPOLY = lbdf.iloc[0].geometry
gmaps = googlemaps.Client(key=os.getenv('GOOGLE_GEOCODING_APIKEY'))

def fix_geolocation_witin(point,addr, poly = LBPOLY):
    new_x = point.x
    new_y = point.y
    updated_geo = False
    if poly.contains(point):
        new_point =  point
    else:
        logger.info(f"Need to fix point {point}")
        geocode_result = gmaps.geocode(addr)
        new_x = geocode_result[0]['geometry']['location']['lng']
        new_y = geocode_result[0]['geometry']['location']['lat']
        new_point = Point(new_x,new_y)
        updated_geo = True
        logger.info(f"Fixed geocoded result for {addr},{new_point}")
    return new_x,new_y,new_point,updated_geo 

#def fix_geolocation_too(x,y):
#    ## This is a hack just based on the bad data patterns we see. Ideally we should geocode based on the addresses
#    new_x = x
#    new_y = y
#    x_pat = re.compile(r'(.*)(\.)(.*)' )
#    x_match = re.match(pat,x)
#    y_match = re.match(pay,y)
#    if(x_match[1] =="188" or x_match[1]=="-188" or x_match[1]=="118"):
#        new_x = "-118"+x_match[2]+x_match[3]
#    if(y_match)

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def correct_scientific_name(c,name_list=None):
    if c in name_list:
        return c
    else:
        for n in name_list:
            if similar(c,n) >= TREE_NAME_MATCH_RATIO:
                logger.debug(f"Updating Treename from {c} to {n}")
                return n
    return c





def generate_tree_names_mapping(infile):
    namedict = {}
    with open(infile,mode='r') as f:
        reader = csv.reader(f)
        namedict = {l[0].title():l[1].title() for l in reader}
    return namedict

def clean_column_name(c):
    trimmed_name  = ' '.join(c.strip().replace('.',' ').replace('#','Number').split())
    if trimmed_name in TREE_STANDARD_MODEL_NAMES.keys():
        return TREE_STANDARD_MODEL_NAMES[' '.join(c.strip().replace('.',' ').replace('#','Number').split())]
    else: 
        return trimmed_name


def normalize_dataframe(df):
    cleaned_column_names = {}
    for c in df.columns:
        cleaned_column_names[c] = clean_column_name(c)
    df.rename(columns=cleaned_column_names,inplace=True)
    return df

def get_treedf_from_sheet(workbook,sheet,rows_to_skip,cols_to_use):
    logger.debug(f"{workbook},{sheet},{rows_to_skip}")
    df = pd.read_excel(workbook,sheet,skiprows=rows_to_skip,usecols=cols_to_use,na_filter=True)
    logger.debug(df)
    df.dropna(axis=0,how='all',inplace=True)
    df = normalize_dataframe(df)
    return df

def process_neighborsvc_treedata(infile):
    logger.debug(f"Processing Neighborhood Services Tree Data -- {infile}")
    ext = os.path.splitext(infile)[-1].lower()
    all_sheets_df = pd.DataFrame()
    if ext == ".xlsx":
        try:
            wb = load_workbook(infile)
            for s in wb.sheetnames:
                logger.debug(f'Sheet name is: {s}')
                df = get_treedf_from_sheet(infile,s,NBRSVC_SKIPROWS,cols_to_use=range(15))
                logger.debug(df)
                logger.debug(f"DF is:\n {df}")
                if all_sheets_df.empty:
                    logger.debug("Combined is None")
                    all_sheets_df = df
                    #all_sheets_df = df.copy()
                else:
                    logger.debug("Concatenating")
                    all_sheets_df = pd.concat([all_sheets_df,df],ignore_index=False)
            logger.debug(">>>>>COMBINED DATAFRAME<<<<<")
            logger.debug(all_sheets_df)
        except Exception as e:
                logger.error(f"Exception occured {e}")
        all_sheets_df.dropna(axis=0,how='all',inplace=True)
        all_sheets_df['Planted By'] = 'Neighborhood Services'
        return all_sheets_df
    else:
        logger.error(f"Filetype {ext} not supported for calfire data")

def process_calfire_treedata(infile):
    quarterly = re.compile('^Q\d+',re.A)
    logger.debug(f"Processing Calfire ---{infile}")
    ext = os.path.splitext(infile)[-1].lower()
    all_sheets_df = pd.DataFrame()
    if ext == ".xlsx":
        logger.info("Processing calfire xlsx file")
        try:
            wb = load_workbook(infile)
            for s in wb.sheetnames:
                logger.debug(f'Sheet name is: {s}')
                match = re.match(quarterly,s)
                if match != None and s not in CALFIRE_SKIP_SHEETS:
                    logger.debug(f'MatchedSheet name is: {s}')
                    df = get_treedf_from_sheet(infile,s,CALFIRE_SKIPROWS,cols_to_use=range(12))
                    logger.debug(f"DF is:\n {df}")
                    if all_sheets_df.empty:
                        logger.debug("Combined is None")
                        all_sheets_df = df
                        #all_sheets_df = df.copy()
                    else:
                        logger.debug("Concatenating")
                        all_sheets_df = pd.concat([all_sheets_df,df],ignore_index=False)
                logger.debug(">>>>>COMBINED DATAFRAME<<<<<")
                logger.debug(all_sheets_df)
        except Exception as e:
                logger.error(f"Exception occured {e}")
        all_sheets_df.dropna(axis=0,how='all',inplace=True)
        all_sheets_df['Planted By'] = 'Conservation Corps of Long Beach'
        return all_sheets_df
    else:
        logger.error(f"Filetype {ext} not supported for calfire data")


def process_tree_data(intype_list,infile_list,outfile,namedict = None):
    combined_df = pd.DataFrame()
    if intype_list == []:
        logger.info("Nothing to Process")
    else:
        if len(intype_list) != len(infile_list):
            raise ValueError("Mismatch in number of sources to be processed and number of input files supplied")
        else:
            logger.debug(f"Processing Inputs: Input types {intype_list}, Input files {infile_list}")
            for t,f in zip(intype_list,infile_list):
                logger.info(f"Processing {t} with file {f}")
                if t == 'CAL_FIRE':
                    combined_df = pd.concat([combined_df,process_calfire_treedata(f)])
                elif t == 'NBR_SVC':
                    combined_df = pd.concat([combined_df,process_neighborsvc_treedata(f)])
                    #raise ValueError("Neighborhood Service File Processing Not currently Implemented")
                elif t == 'OFC_SUSTAIN':
                    raise ValueError("Office of Sustainability File Processing Not currently Implemented")
                else:
                    raise ValueError("Unknown type for processing")
                logger.info(f"Processed {f}, current DF shape {combined_df.shape}")
            logger.info(f"Writing {combined_df.shape[0]} rows and {combined_df.shape[1]} columns to {outfile}")
            if namedict != None:
                logger.info("Correcting Scientific Names")
                combined_df['Scientific Name'] = combined_df.apply(lambda r: correct_scientific_name(r['Scientific Name'].title(),namedict.keys()),axis=1)
                logger.info("Filling Common Names")
                combined_df['Common Name'] = combined_df.apply(lambda r: namedict.get(r['Scientific Name'].title(),"Unknown") if pd.isnull(r['Common Name']) else r["Common Name"].title(), axis = 1)
            print(combined_df['Common Name'])
            combined_gdf = gpd.GeoDataFrame(combined_df, geometry=gpd.points_from_xy(combined_df['X Coordinate'], combined_df['Y Coordinate']))
            #combined_gdf['x'],combined_gdf['y'],combined_gdf['geometry'], combined_gdf['did_update_geo'] = combined_gdf.apply(lambda r: fix_geolocation_witin(r['geometry'],r['Street Address']+',Long Beach, CA'),axis=1,result_type='expand')
            combined_gdf[["X Coordinate","Y Coordinate","geometry",'did_update_geo']]= combined_gdf.apply(lambda r: fix_geolocation_witin(r['geometry'],r['Street Address']+',Long Beach, CA'),axis=1,result_type='expand')
            logger.info(combined_df)
            #combined_df.to_csv(outfile,index=False)
            combined_gdf.to_file(outfile,driver="GeoJSON" )

if __name__ == "__main__":
    logger.info("Inside Main")
    parser = argparse.ArgumentParser(description = 'Ingests internal tree datafiles and spits out standard tree data model')
    parser.add_argument('-t/--intype', dest = 'intypes', action = 'append',choices = ['CAL_FIRE','NBR_SVC','OFC_SUSTAIN'], required=True, help = "Source type of data file to be processed ")
    parser.add_argument('-i/--infile', dest = 'infiles',  action = 'append',required=True, help = "Input file name associated with the source type. This has to follow the same order as source type for meaningful results")
    parser.add_argument('-o/--outfile', dest = 'outfile',  action = 'store',required=True, help = "Output file name associated with the source type.")
    parser.add_argument('-n/--namefile',dest = 'namefile', action = 'store', required = False, help = 'Ingests tree name mapping file')
    args = parser.parse_args()
    namedict = generate_tree_names_mapping(args.namefile)
    process_tree_data(args.intypes,args.infiles,args.outfile,namedict)
